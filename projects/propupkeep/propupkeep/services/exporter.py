from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from propupkeep.models.issue import IssueReport


EXPORT_COLUMNS = [
    "issue_id",
    "created_at",
    "updated_at",
    "status",
    "source",
    "property_name",
    "building",
    "unit",
    "area",
    "issue",
    "urgency",
    "category",
    "recommended_action",
    "reported_observation",
    "recipients",
    "followup_questions",
    "confidence",
    "location_conflict",
    "image_filename",
    "extracted_entities",
    "comment_count",
    "latest_comment",
    "latest_comment_at",
]


def export_issues_to_excel_bytes(issues: list[IssueReport]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Issue Reports"
    worksheet.append(EXPORT_COLUMNS)

    for issue in issues:
        latest_comment = _get_latest_comment(issue)
        worksheet.append(
            [
                issue.report_id,
                _to_excel_datetime(issue.created_at),
                _to_excel_datetime(issue.updated_at),
                issue.status.value,
                issue.source.value,
                issue.property_name,
                issue.building,
                issue.unit_number,
                issue.area or "",
                issue.issue,
                issue.urgency.value,
                issue.category.value,
                issue.recommended_action,
                issue.reported_observation,
                ", ".join(issue.recipients),
                ", ".join(issue.followup_questions),
                json.dumps(issue.confidence.model_dump(mode="json"), ensure_ascii=True),
                getattr(issue, "location_conflict", "") or "",
                issue.image_filename or "",
                json.dumps(issue.extracted_entities.model_dump(mode="json"), ensure_ascii=True),
                len(issue.comments),
                latest_comment["message"],
                latest_comment["created_at"],
            ]
        )

    _apply_column_widths(worksheet)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _to_excel_datetime(value: datetime | None) -> str:
    if not value:
        return ""
    if value.tzinfo is None:
        as_utc = value.replace(tzinfo=timezone.utc)
    else:
        as_utc = value.astimezone(timezone.utc)
    return as_utc.strftime("%Y-%m-%d %H:%M:%S UTC")


def _get_latest_comment(issue: IssueReport) -> dict[str, str]:
    if not issue.comments:
        return {"message": "", "created_at": ""}
    latest = max(issue.comments, key=lambda comment: comment.created_at)
    return {
        "message": latest.message,
        "created_at": _to_excel_datetime(latest.created_at),
    }


def _apply_column_widths(worksheet) -> None:
    max_width = 60
    min_width = 12
    for idx, column_name in enumerate(EXPORT_COLUMNS, start=1):
        longest = len(column_name)
        for row_idx in range(2, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_idx, column=idx).value
            if cell_value is None:
                continue
            cell_length = len(str(cell_value))
            if cell_length > longest:
                longest = cell_length

        worksheet.column_dimensions[get_column_letter(idx)].width = min(
            max(longest + 2, min_width),
            max_width,
        )
